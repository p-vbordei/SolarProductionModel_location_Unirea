"""
Export intraday forecast CSV files to Excel format

IMPORTANT: All timestamps in Excel output are in CET/CEST (Europe/Berlin) timezone
"""
import pandas as pd
import os
from datetime import datetime
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys

def export_forecast_to_excel(output_path=None):
    """
    Export the latest intraday forecast CSV files to Excel with weather data

    Args:
        output_path: Path for Excel file (optional, defaults to data_output/intraday)
    """
    try:
        # Set data directory
        data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data_output', 'intraday')

        # Find latest weather data file
        # PRIORITY: Use processed_weather_local.csv if available (has correct local solar calculations)
        # FALLBACK: Use raw_weather_data.csv if processed not available
        weather_data = None
        param_dirs = [d for d in os.listdir(data_dir) if d.startswith('parameters-') and os.path.isdir(os.path.join(data_dir, d))]
        if param_dirs:
            # Get the most recent parameters directory
            param_dirs.sort()
            latest_param_dir = param_dirs[-1]

            # Try processed weather data first (preferred - has local solar calculations)
            processed_weather_file = os.path.join(data_dir, latest_param_dir, 'processed_weather_local.csv')
            raw_weather_file = os.path.join(data_dir, latest_param_dir, 'raw_weather_data.csv')

            if os.path.exists(processed_weather_file):
                print(f"✅ Loading processed weather data (local solar calculations): {processed_weather_file}")
                # Read processed weather data with first column as timestamp
                weather_data = pd.read_csv(processed_weather_file, index_col=0, parse_dates=True)
                # Processed weather is already in CET/CEST and naive (no timezone)
                print(f"Processed weather data loaded: {len(weather_data)} rows")
            elif os.path.exists(raw_weather_file):
                print(f"⚠️  Loading raw weather data (UTC solar timing): {raw_weather_file}")
                print("   Note: Processed weather not found. Using raw weather may have timezone offset issues.")
                # Read weather data with first column as timestamp
                weather_data = pd.read_csv(raw_weather_file, index_col=0, parse_dates=True)
                # Convert timezone to CET for merging
                if weather_data.index.tz is not None:
                    weather_data.index = weather_data.index.tz_convert('Europe/Berlin')
                print(f"Raw weather data loaded: {len(weather_data)} rows")

        # Find latest CSV files
        csv_files = [f for f in os.listdir(data_dir) if f.endswith('.csv') and 'cm_forecast_intraday' in f]
        
        # Separate 15min and 1hour files
        files_15min = [f for f in csv_files if '15min' in f]
        files_1hour = [f for f in csv_files if '1hour' in f]
        
        if not files_15min or not files_1hour:
            print("Error: Could not find forecast CSV files")
            print(f"Looking in directory: {data_dir}")
            print(f"Found {len(files_15min)} 15-minute files and {len(files_1hour)} hourly files")
            return None
        
        # Sort by modification time to get the truly latest files
        files_15min.sort(key=lambda x: os.path.getmtime(os.path.join(data_dir, x)))
        files_1hour.sort(key=lambda x: os.path.getmtime(os.path.join(data_dir, x)))
        
        # Use latest files
        latest_15min = files_15min[-1]
        latest_1hour = files_1hour[-1]
        
        print(f"Loading 15-minute data: {latest_15min}")
        print(f"Loading hourly data: {latest_1hour}")
        
        # Read CSV files (skip comment lines starting with #)
        df_15min = pd.read_csv(os.path.join(data_dir, latest_15min), comment='#')
        df_1hour = pd.read_csv(os.path.join(data_dir, latest_1hour), comment='#')

        # Merge weather data if available
        if weather_data is not None:
            # Prepare weather data for merging
            weather_cet = weather_data.copy()

            # Only process timezone for raw weather data
            if weather_cet.index.tz is not None:
                # Convert weather data to CET/CEST and make naive (strip timezone)
                weather_cet.index = weather_cet.index.tz_convert('Europe/Berlin')
                # Strip timezone info to match naive forecast timestamps
                weather_cet.index = weather_cet.index.tz_localize(None)

                # Handle edge case: forecast may start before weather data (only for raw weather)
                # Extend weather backwards by creating rows at earlier timestamps
                first_weather_time = weather_cet.index.min()
                # Create hourly timestamps going back 2 hours before first weather timestamp
                earlier_times = pd.date_range(
                    end=first_weather_time - pd.Timedelta(minutes=15),
                    periods=8,  # 2 hours at 15-min resolution
                    freq='15min'
                )
                # Create dataframe with first weather values for earlier times
                first_weather_values = weather_cet.iloc[0:1]
                earlier_weather = pd.DataFrame(
                    [first_weather_values.values[0]] * len(earlier_times),
                    index=earlier_times,
                    columns=weather_cet.columns
                )
                # Combine earlier weather with actual weather
                weather_cet = pd.concat([earlier_weather, weather_cet]).sort_index()
            # Processed weather is already in CET/CEST and naive - no processing needed

            # For 15-minute data
            if 'timestamp' in df_15min.columns or df_15min.columns[0] in ['', 'Unnamed: 0']:
                # Get timestamp column
                ts_col = 'timestamp' if 'timestamp' in df_15min.columns else df_15min.columns[0]
                # Parse timestamps - they are already in local time (naive)
                df_15min['timestamp_parsed'] = pd.to_datetime(df_15min[ts_col])

                # Prepare weather dataframe for merge
                weather_cols = ['temperature', 'ghi', 'dni', 'dhi', 'wind_speed', 'cloud_cover', 'humidity']
                available_weather_cols = [c for c in weather_cols if c in weather_cet.columns]
                weather_merge = weather_cet[available_weather_cols].reset_index()
                weather_merge.columns = ['timestamp_parsed'] + available_weather_cols

                # Use merge_asof to match nearest timestamps (handles small timezone mismatches)
                df_15min = pd.merge_asof(
                    df_15min.sort_values('timestamp_parsed'),
                    weather_merge.sort_values('timestamp_parsed'),
                    on='timestamp_parsed',
                    direction='nearest',
                    tolerance=pd.Timedelta('1 hour')
                )
                df_15min = df_15min.drop(columns=['timestamp_parsed'])
                print(f"Added {len(available_weather_cols)} weather columns to 15-minute data")

            # For hourly data - aggregate weather data to hourly
            if 'timestamp' in df_1hour.columns or df_1hour.columns[0] in ['', 'Unnamed: 0']:
                # Get timestamp column
                ts_col = 'timestamp' if 'timestamp' in df_1hour.columns else df_1hour.columns[0]
                # Parse timestamps - they are already in local time (naive)
                df_1hour['timestamp_parsed'] = pd.to_datetime(df_1hour[ts_col])

                # Aggregate weather data to hourly (mean) - index is already naive CET
                weather_hourly = weather_cet.resample('h').mean()

                # Prepare hourly weather for merge
                weather_cols = ['temperature', 'ghi', 'dni', 'dhi', 'wind_speed', 'cloud_cover', 'humidity']
                available_weather_cols = [c for c in weather_cols if c in weather_hourly.columns]
                weather_hourly_merge = weather_hourly[available_weather_cols].reset_index()
                weather_hourly_merge.columns = ['timestamp_parsed'] + available_weather_cols

                # Use merge_asof to match nearest timestamps (handles small timezone mismatches)
                df_1hour = pd.merge_asof(
                    df_1hour.sort_values('timestamp_parsed'),
                    weather_hourly_merge.sort_values('timestamp_parsed'),
                    on='timestamp_parsed',
                    direction='nearest',
                    tolerance=pd.Timedelta('1 hour')
                )
                df_1hour = df_1hour.drop(columns=['timestamp_parsed'])
                print(f"Added {len(available_weather_cols)} weather columns to hourly data")

        # Convert MWh to kWh (multiply by 1000)
        energy_columns = ['energy_mwh', 'energy_q10_mwh', 'energy_q25_mwh', 'energy_q50_mwh', 'energy_q75_mwh', 'energy_q90_mwh']
        for col in energy_columns:
            if col in df_15min.columns:
                df_15min[col] = df_15min[col] * 1000
            if col in df_1hour.columns:
                df_1hour[col] = df_1hour[col] * 1000

        # Convert power from MW to kW if needed
        if 'power_mw' in df_1hour.columns:
            df_1hour['power_kw'] = df_1hour['power_mw'] * 1000
        elif 'production_mw' in df_1hour.columns:
            df_1hour['power_kw'] = df_1hour['production_mw'] * 1000
        elif 'production_kw' in df_1hour.columns:
            df_1hour['power_kw'] = df_1hour['production_kw']

        if 'power_mw' in df_15min.columns:
            df_15min['power_kw'] = df_15min['power_mw'] * 1000
        elif 'production_mw' in df_15min.columns:
            df_15min['power_kw'] = df_15min['production_mw'] * 1000
        elif 'production_kw' in df_15min.columns:
            df_15min['power_kw'] = df_15min['production_kw']

        # Convert quartile power values from MW to kW for both 15-minute and hourly data
        quartile_columns = ['q10', 'q25', 'q50', 'q75', 'q90']
        for q_col in quartile_columns:
            if q_col in df_15min.columns:
                # Convert existing quartile column from MW to kW (in place)
                df_15min[q_col] = df_15min[q_col] * 1000
            if q_col in df_1hour.columns:
                # Convert existing quartile column from MW to kW (in place)
                df_1hour[q_col] = df_1hour[q_col] * 1000

        # Rename columns from MWh to kWh
        rename_map = {
            'energy_mwh': 'energy_kwh',
            'energy_q10_mwh': 'energy_q10_kwh',
            'energy_q25_mwh': 'energy_q25_kwh',
            'energy_q50_mwh': 'energy_q50_kwh',
            'energy_q75_mwh': 'energy_q75_kwh',
            'energy_q90_mwh': 'energy_q90_kwh'
        }
        df_15min = df_15min.rename(columns=rename_map)
        df_1hour = df_1hour.rename(columns=rename_map)

        # Fix column names if timestamp is unnamed
        if 'Unnamed: 0' in df_15min.columns:
            df_15min = df_15min.rename(columns={'Unnamed: 0': 'timestamp'})
        elif df_15min.columns[0] == '' or pd.isna(df_15min.columns[0]):
            df_15min = df_15min.rename(columns={df_15min.columns[0]: 'timestamp'})
            
        if 'Unnamed: 0' in df_1hour.columns:
            df_1hour = df_1hour.rename(columns={'Unnamed: 0': 'timestamp'})
        elif df_1hour.columns[0] == '' or pd.isna(df_1hour.columns[0]):
            df_1hour = df_1hour.rename(columns={df_1hour.columns[0]: 'timestamp'})
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Add header
        ws_summary['A1'] = "CRIPVSOL ENERGY SRL - Unirea Solar Forecast Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A3'] = f"Generated: {datetime.now(pytz.timezone('Europe/Berlin')).strftime('%Y-%m-%d %H:%M:%S CET')}"

        # Add location info
        ws_summary['A5'] = "Location Information"
        ws_summary['A5'].font = Font(bold=True)
        ws_summary['A6'] = "Plant Name:"
        ws_summary['B6'] = "CRIPVSOL ENERGY SRL"
        ws_summary['A7'] = "Capacity:"
        ws_summary['B7'] = "2.9 MW AC (2.916 MW DC)"
        ws_summary['A8'] = "Location:"
        ws_summary['B8'] = "45.1116, 27.7846 (Unirea, Brăila)"
        ws_summary['A9'] = "Timezone:"
        ws_summary['B9'] = "CET (Central European Time)"
        ws_summary['A10'] = "Panels:"
        ws_summary['B10'] = "Canadian Solar CS6W-540 540W x 5,400"
        ws_summary['A11'] = "Inverters:"
        ws_summary['B11'] = "Huawei SUN2000-215KTL-H0 185kW x 16"
        
        # Add forecast summary
        ws_summary['A13'] = "Forecast Summary"
        ws_summary['A13'].font = Font(bold=True)
        ws_summary['A14'] = "Forecast Period:"
        ws_summary['B14'] = f"{df_1hour['timestamp'].iloc[0]} to {df_1hour['timestamp'].iloc[-1]}"
        ws_summary['A15'] = "Duration:"
        ws_summary['B15'] = "168 hours (7 days)"
        ws_summary['A16'] = "Peak Power:"
        ws_summary['B16'] = f"{df_1hour['power_kw'].max():.1f} kW"
        ws_summary['A17'] = "Total Energy:"
        ws_summary['B17'] = f"{df_1hour['energy_kwh'].sum():.1f} kWh"
        ws_summary['A18'] = "Average Power:"
        ws_summary['B18'] = f"{df_1hour['power_kw'].mean():.1f} kW"
        
        # Format summary cells
        for row in ws_summary.iter_rows(min_row=1, max_row=18, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(vertical='center')
        
        # Add hourly sheet
        ws_hourly = wb.create_sheet("Hourly Forecast", 1)

        # Select columns for hourly sheet in exact order (including weather data)
        hourly_columns = ['timestamp', 'YEAR', 'MONTH', 'DAY', 'HOUR_START', 'HOUR_END', 'DAY_END', 'interval',
                         'power_kw', 'energy_kwh', 'q10', 'q25', 'q50', 'q75', 'q90',
                         'temperature', 'ghi', 'dni', 'dhi', 'wind_speed', 'cloud_cover', 'humidity']
        hourly_columns_filtered = [col for col in hourly_columns if col in df_1hour.columns]
        df_1hour_display = df_1hour[hourly_columns_filtered]

        # Add header row with formatting
        headers = list(df_1hour_display.columns)
        ws_hourly.append(headers)
        
        # Format header
        for cell in ws_hourly[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for index, row in df_1hour_display.iterrows():
            ws_hourly.append(row.tolist())
        
        # Format columns
        for column in ws_hourly.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_hourly.column_dimensions[column_letter].width = adjusted_width
        
        # Add 15-minute sheet
        ws_15min = wb.create_sheet("15 Minutes", 2)
        
        # Rename power_kw to production_kw to match specification
        if 'power_kw' in df_15min.columns:
            df_15min = df_15min.rename(columns={'power_kw': 'production_kw'})

        # Prepare 15-minute data with exact column order as specified (including weather data)
        columns_to_include = [
            # Time components
            'timestamp', 'YEAR', 'MONTH', 'DAY', 'HOUR_START', 'MINUTE_START', 'HOUR_END', 'MINUTE_END', 'DAY_END', 'interval',
            # Power/Energy
            'production_kw', 'energy_kwh',
            # Quantiles
            'q10', 'energy_q10_kwh',
            'q25', 'energy_q25_kwh',
            'q50', 'energy_q50_kwh',
            'q75', 'energy_q75_kwh',
            'q90', 'energy_q90_kwh',
            # Weather data
            'temperature', 'ghi', 'dni', 'dhi', 'wind_speed', 'cloud_cover', 'humidity'
        ]

        # Filter to only existing columns
        columns_to_include = [col for col in columns_to_include if col in df_15min.columns]
        df_15min_ordered = df_15min[columns_to_include]
        
        # Add headers
        headers_15min = list(df_15min_ordered.columns)
        ws_15min.append(headers_15min)
        
        # Format header
        for cell in ws_15min[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for index, row in df_15min_ordered.iterrows():
            ws_15min.append(row.tolist())
        
        # Format columns
        for column in ws_15min.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_15min.column_dimensions[column_letter].width = adjusted_width
        
        # Generate output filename
        if output_path is None:
            timestamp = datetime.now(pytz.timezone('Europe/Berlin')).strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(data_dir, f'cripvsol_unirea_forecast_{timestamp}.xlsx')
        
        # Save workbook
        wb.save(output_path)
        print(f"\n✅ Excel file created: {output_path}")
        
        # Display summary
        print("\n📊 Forecast Summary:")
        print(f"   Location: CRIPVSOL ENERGY SRL - Unirea (2.9 MW)")
        print(f"   Period: 7 days ({len(df_1hour)} hours)")
        print(f"   Peak Power: {df_1hour['power_kw'].max():.1f} kW")
        print(f"   Total Energy: {df_1hour['energy_kwh'].sum():.1f} kWh")
        print(f"   Data Resolution: 15-minute and hourly")
        
        return output_path
    
    except Exception as e:
        print(f"Error in export_forecast_to_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    result = export_forecast_to_excel()
    if result is None:
        sys.exit(1)
    sys.exit(0)